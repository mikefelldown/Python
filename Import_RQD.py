#-------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      mfell
#
# Created:     23/04/2015
# Copyright:   (c) Canadian Malartic Corporation 2015
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import pandas
import openpyxl
import os

os.chdir("C:\GIS Stuff\Upper Beaver\Geotech 2")
file_list = os.listdir(os.getcwd())

def findEOH(df):
    for d in df.iterrows():
        if 'EOH' in d[1].values:
            return d[0]
    return None


for xl in file_list:
    out_file = openpyxl.load_workbook(r'C:\GIS Stuff\Upper Beaver\Template_2012.xlsx')
    ws = out_file.get_sheet_by_name('RQD')
    ws2 = out_file.get_sheet_by_name('Perso_Advanced RQD')
    head =pandas.io.excel.read_excel("file:\\\\localhost\\" + os.getcwd() + "\\" + xl, "INPUT", parse_cols = 2)
    skip = head[head['SRK IRMR LOGGING WORKSHEET']=="Drillhole ID"].index.values[0] + 1
    data = pandas.io.excel.read_excel("file:\\\\localhost\\" + os.getcwd() + "\\" + xl, "INPUT", skiprows = skip, header = 0)
    project = "UpperBeaver"
    hole = head['Unnamed: 1'][2]
    EOH = findEOH(data)
    data = data[:EOH]
    data = data[data["From (m)"].notnull()]
    for i in data.iterrows():
        next_row = ws.get_highest_row() +1
        ws.cell(row = next_row, column = 1).value = ws2.cell(row = next_row, column = 1).value = project
        ws.cell(row = next_row, column = 2).value = ws2.cell(row = next_row, column = 2).value = i[1]['Drillhole ID'] if i[1]['Drillhole ID'] is not None else hole
        ws.cell(row = next_row, column = 3).value = ws2.cell(row = next_row, column = 3).value = i[1]['From (m)']
        ws.cell(row = next_row, column = 4).value = ws2.cell(row = next_row, column = 4).value = i[1]['To (m)']
        ws.cell(row = next_row, column = 7).value = i[1]['Length (m)']
        ws.cell(row = next_row, column = 9).value = i[1]['Length (m).1']
        ws2.cell(row = next_row, column = 5).value = i[1][5]
        ws2.cell(row = next_row, column = 6).value = i[1][6]
        ws2.cell(row = next_row, column = 7).value = i[1][7]
        ws2.cell(row = next_row, column = 8).value = i[1][8]
        ws2.cell(row = next_row, column = 9).value = i[1][9]
        ws2.cell(row = next_row, column = 10).value = i[1][10]
        ws2.cell(row = next_row, column = 11).value = i[1][11]
        ws2.cell(row = next_row, column = 12).value = i[1][12]
        ws2.cell(row = next_row, column = 13).value = i[1][13]
        ws2.cell(row = next_row, column = 14).value = i[1][14]
        ws2.cell(row = next_row, column = 15).value = i[1][15]
        ws2.cell(row = next_row, column = 16).value = i[1][16]
        ws2.cell(row = next_row, column = 17).value = i[1][17]
        ws2.cell(row = next_row, column = 18).value = i[1][18]
        ws2.cell(row = next_row, column = 19).value = i[1][19]
        ws2.cell(row = next_row, column = 20).value = i[1][20]
        ws2.cell(row = next_row, column = 21).value = i[1][21]
        ws2.cell(row = next_row, column = 22).value = i[1][22]
        ws2.cell(row = next_row, column = 23).value = i[1][23]
        ws2.cell(row = next_row, column = 24).value = i[1][24]
        ws2.cell(row = next_row, column = 25).value = i[1][25]
        ws2.cell(row = next_row, column = 26).value = i[1][26]
        ws2.cell(row = next_row, column = 27).value = i[1][27]
        ws2.cell(row = next_row, column = 28).value = i[1][28]
        ws2.cell(row = next_row, column = 29).value = i[1][29]
        ws2.cell(row = next_row, column = 30).value = i[1][30]
        ws2.cell(row = next_row, column = 31).value = i[1][31]
        ws2.cell(row = next_row, column = 32).value = i[1][32]
        ws2.cell(row = next_row, column = 33).value = i[1][33]
        ws2.cell(row = next_row, column = 34).value = i[1][34]
        ws2.cell(row = next_row, column = 35).value = i[1][35]
        ws2.cell(row = next_row, column = 36).value = i[1][36]
        ws2.cell(row = next_row, column = 37).value = i[1][37]
        ws2.cell(row = next_row, column = 38).value = i[1][38]





    out_file.save(r'C:\GIS Stuff\Upper Beaver\Template_2012.xlsx')



##def main():
##    pass
##
##if __name__ == '__main__':
##    main()
